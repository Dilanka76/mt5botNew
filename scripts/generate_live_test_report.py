"""Generates reports/live_test/live_test_report.xlsx — a daily Excel report
tracking live/shadow testing of the currently-deployed strategy across both
timeframe variants of one account pair. Defaults to demo1_m1 (M1) and
demo1_m3 (M3); pass --m1-account/--m3-account/--output-name to point it at
a different pair (e.g. demo2_m1/demo2_m3), generalized 2026-08-27 at the
user's request for a separate report covering demo2.

    python scripts/generate_live_test_report.py
    python scripts/generate_live_test_report.py --m1-account demo2_m1 --m3-account demo2_m3 --output-name demo2_report.xlsx

The "Strategy Rule Tracking" tab's content genuinely differs by account:
each account's config.strategy_variant is read fresh at report-run time
and used to pick the right rule breakdown (see build_rule_tracking_swap_adx
vs build_rule_tracking_simple_swap below) -- demo1's dual_cross_confirmed_
swap_adx has an ADX-gated, 2-candle-debounced swap and stop-loss
tightening; demo2's dual_cross_confirmed_swap has none of those (see
bot/strategy/state_machine_dual_cross_confirmed_swap.py's module
docstring). Reusing demo1's rule-tracking numbers for demo2 would be
actively misleading, not just cosmetically wrong.

No date args: always covers TESTING_START_UTC (below) through now, and
always rewrites its output file from scratch — idempotent by construction
(nothing is appended to an existing file, so re-running never duplicates
a row).

TESTING_START_UTC is a fixed cutoff, NOT "earliest decisions.jsonl
entry" — explicit user request 2026-08-26: don't mix in trades from
earlier, retired strategy variants (this project has been through
several: dual_cross, cross_confirmed, dual_cross_tight_exit,
dual_cross_confirmed_swap_adx, dual_cross_confirmed_adx_m15, back to
dual_cross_confirmed_swap_adx...) — this report tracks the
CURRENTLY-deployed strategy's live/shadow testing specifically, starting
fresh from today. Update the constant if the user ever wants to reset
the tracking window again (e.g. after a major strategy change).

Two sources, joined per trade:
  - MT5 deal history (mt5.history_deals_get, via bot.analytics.
    get_closed_trades_range) — authoritative entry/exit price, P/L
    (swap+commission-adjusted), open/close time. Only real orders
    (execution.mode demo_execute/live_execute) show up here; a
    shadow-mode account has no MT5 deals at all and contributes nothing
    to the $-denominated tabs (Dashboard/Summary/Detail/Attribution) —
    only to Strategy Rule Tracking, which is built from decisions.jsonl
    alone.
  - logs/<account>/decisions.jsonl (+ rotated .1..5 backups) — the bot's
    own reasoning: entry gap size, close-reason category (MT5 can't tell
    a bot-driven stop-loss close apart from a swap-reversal close; both
    are manual closes tagged identically by trade_executor.close_position,
    see classify_exit_reason's docstring in bot/analytics.py), session
    blocks, ADX blocks, pending swap/gap setups.

Designed to also run unattended once daily via Windows Task Scheduler at
23:55 Asia/Colombo (see setup_live_test_report_task.ps1) — read-only
against MT5 (only queries history, never places/closes a trade) and safe
to re-run any number of times.
"""
from __future__ import annotations

import argparse
import glob
import json
import re
import sys
from collections import Counter
from datetime import date as date_cls, datetime, time as dt_time, timedelta, timezone
from pathlib import Path
from statistics import mean

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bot.analytics import COLOMBO, get_closed_trades_range, mt5_utc_offset
from bot.config import PROJECT_ROOT, SessionWindow, load_config

# The report's calendar day runs 03:30-to-03:30 Sri Lanka time -- see
# trading_day()'s own docstring below for the full history/reasoning.
DAY_BOUNDARY = timedelta(hours=3, minutes=30)

# Fixed cutoff, not "earliest decisions.jsonl entry" -- see module
# docstring. 03:30 Sri Lanka time on 2026-08-25, the same day-boundary
# convention as trading_day() below (explicit user instruction
# 2026-08-27: "the 25th is the start date... finalize the analyze report
# with this daily time") -- so the very first day this report ever
# counts is itself a complete 03:30-to-03:30 trading day under the
# current convention, not a partial one under some earlier convention.
TESTING_START_UTC = datetime(2026, 8, 25, 3, 30, tzinfo=COLOMBO).astimezone(timezone.utc)


def trading_day(dt: datetime) -> date_cls:
    """This report's calendar day runs 03:30-to-03:30 in Sri Lanka time
    (Asia/Colombo, UTC+5:30) -- explicit user instruction 2026-08-27,
    replacing the earlier midnight-to-midnight MT5 APP-time convention
    (which matched the MT5 app's own "Today" view but is no longer what
    the user wants here). Anchored by subtracting the boundary offset
    before taking the date: a moment at or after 03:30 Colombo belongs to
    THAT calendar date's trading day; anything before 03:30 still belongs
    to the PREVIOUS date's trading day (it hasn't rolled over yet)."""
    return (dt.astimezone(COLOMBO) - DAY_BOUNDARY).date()
from bot.mt5_connector import MT5Connector

ENTRY_PAIR_WINDOW_SECONDS = 300  # max gap between a trade_entered log line and the MT5 fill it caused

GAP_RE = re.compile(r"gap=(-?\d+\.?\d*)")

# -------------------------------------------------------------------------
# decisions.jsonl reading
# -------------------------------------------------------------------------


def _rotated_paths(log_dir: Path) -> list[Path]:
    """decisions.jsonl (current, MOST recent lines) plus RotatingFileHandler
    backups decisions.jsonl.1 (next-most-recent) .. .N (oldest) — ordered
    oldest-file-first so concatenating their lines gives full chronological
    order."""
    base = log_dir / "decisions.jsonl"
    numbered = sorted(
        glob.glob(str(log_dir / "decisions.jsonl.*")),
        key=lambda p: -int(p.rsplit(".", 1)[1]) if p.rsplit(".", 1)[1].isdigit() else 0,
    )
    paths = [Path(p) for p in numbered if p.rsplit(".", 1)[1].isdigit()]
    if base.exists():
        paths.append(base)
    return paths


def read_decisions(account: str) -> list[dict]:
    log_dir = PROJECT_ROOT / "logs" / account
    entries: list[dict] = []
    for path in _rotated_paths(log_dir):
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                entry = json.loads(line)
                entry["_ts"] = datetime.fromisoformat(entry["timestamp"])
                entries.append(entry)
            except (json.JSONDecodeError, KeyError, ValueError):
                continue
    entries.sort(key=lambda e: e["_ts"])
    return entries


# -------------------------------------------------------------------------
# Joining MT5 truth with decisions.jsonl reasoning
# -------------------------------------------------------------------------


def _extract_gap(reason: str) -> float | None:
    m = GAP_RE.search(reason or "")
    return float(m.group(1)) if m else None


def build_trade_records(account: str, timeframe: str, decisions: list[dict], mt5_trades: list[dict]) -> list[dict]:
    """One row per closed MT5 trade, enriched with this bot's own
    close-reason category and entry gap size pulled from decisions.jsonl."""
    exits_by_ticket: dict[int, dict] = {}
    for e in decisions:
        if e.get("action") in ("trade_exited", "trade_closed_tp") and e.get("ticket") is not None:
            category = e.get("category") if e["action"] == "trade_exited" else "take_profit"
            exits_by_ticket[e["ticket"]] = {"category": category, "reason": e.get("reason", "")}

    entries = [e for e in decisions if e.get("action") == "trade_entered"]
    used_entry_idx: set[int] = set()

    records = []
    for t in mt5_trades:
        exit_info = exits_by_ticket.get(t["ticket"])
        if exit_info:
            category = exit_info["category"] or "unknown"
        else:
            # No decisions.jsonl match (log rotated away, or a
            # startup-reconciled position) -- fall back to MT5's own
            # coarser classification.
            mt5_reason = t["exit_reason"]
            if mt5_reason == "Take Profit":
                category = "take_profit"
            elif mt5_reason == "Stop Loss":
                category = "stop_loss"
            else:
                category = "bot_close_unclassified"

        # Pair to the trade_entered log line closest in time (same
        # direction, within ENTRY_PAIR_WINDOW_SECONDS) to recover the
        # entry gap size -- swap re-entries never log a gap (see
        # state_machine_dual_cross_confirmed_swap_adx.py's module
        # docstring: the gap+EMA5-pullback rule only applies to flat
        # entries), so this legitimately comes back None for those.
        best_idx, best_delta = None, None
        for i, e in enumerate(entries):
            if i in used_entry_idx or e.get("direction") != t["direction"]:
                continue
            delta = abs((e["_ts"] - t["entry_time"].astimezone(timezone.utc)).total_seconds())
            if delta <= ENTRY_PAIR_WINDOW_SECONDS and (best_delta is None or delta < best_delta):
                best_idx, best_delta = i, delta
        gap = None
        if best_idx is not None:
            used_entry_idx.add(best_idx)
            gap = _extract_gap(entries[best_idx].get("reason", ""))

        profit = t["profit"]
        outcome = "WIN" if profit > 0 else ("LOSS" if profit < 0 else "BREAKEVEN")

        records.append({
            "account": account,
            "timeframe": timeframe,
            "ticket": t["ticket"],
            "direction": t["direction"],
            "entry_time": t["entry_time"],
            "exit_time": t["exit_time"],
            "entry_price": t["entry_price"],
            "exit_price": t["exit_price"],
            "volume": t["volume"],
            "profit": profit,
            "outcome": outcome,
            "close_category": category,
            "entry_gap": gap,
        })

    records.sort(key=lambda r: r["exit_time"])
    return records


CATEGORY_LABELS = {
    "stop_loss": "Stop-Loss",
    "breakeven": "Breakeven-Stop",
    "swapped_confirmed_reversal": "Swap Reversal",
    "swapped_reversal": "Swap Reversal (immediate)",
    "take_profit": "Take-Profit",
    "closed_by_concurrent_validation": "Concurrent Validation Close",
    "bot_close_unclassified": "Bot Close (uncategorized)",
    "unknown": "Unknown",
}


def category_label(category: str) -> str:
    return CATEGORY_LABELS.get(category, category)


# -------------------------------------------------------------------------
# Strategy Rule Tracking (Tab 5) -- built from decisions.jsonl alone
# -------------------------------------------------------------------------


def _histogram(values: list[float], bin_width: float = 1.0) -> list[tuple[str, int]]:
    if not values:
        return []
    max_v = max(values)
    n_bins = int(max_v // bin_width) + 1
    counts = [0] * max(n_bins, 1)
    for v in values:
        idx = min(int(v // bin_width), len(counts) - 1)
        counts[idx] += 1
    return [(f"${i * bin_width:.0f}-${(i + 1) * bin_width:.0f}", c) for i, c in enumerate(counts)]


def _session_gap_label(t: datetime, windows: list[SessionWindow]) -> str:
    """Which inter-window gap a blocked-by-session timestamp fell in --
    the answer to "which of the configured windows was inactive", phrased
    as the gap between the two windows bracketing it (there can be more
    or fewer than 3 depending on the account's current config)."""
    if not windows:
        return "No session windows configured"
    colombo_t = t.astimezone(COLOMBO).time()
    ordered = sorted(windows, key=lambda w: w.start)
    n = len(ordered)
    for i in range(n):
        gap_start = ordered[i].end
        gap_end = ordered[(i + 1) % n].start
        if _time_in_range(colombo_t, gap_start, gap_end):
            return f"Between {ordered[i].start}-{ordered[i].end} and {ordered[(i + 1) % n].start}-{ordered[(i + 1) % n].end}"
    return "Unclassified (overlapping windows?)"


def _time_in_range(t: dt_time, start_str: str, end_str: str) -> bool:
    sh, sm = start_str.split(":")
    eh, em = end_str.split(":")
    start, end = dt_time(int(sh), int(sm)), dt_time(int(eh), int(em))
    if start <= end:
        return start <= t <= end
    return t >= start or t <= end  # wraps past midnight


def build_rule_tracking_swap_adx(account: str, decisions: list[dict], records: list[dict], sessions: list[SessionWindow]) -> dict:
    """For strategy_variant=dual_cross_confirmed_swap_adx (demo1_m1/
    demo1_m3): ADX-gated, 2-candle-debounced swap; stop-loss tightening
    on the first opposing candle. See build_rule_tracking_simple_swap for
    demo2's genuinely different, simpler engine -- do NOT reuse this
    function's numbers for an account running a different variant."""
    entries = [e for e in decisions if e.get("action") == "trade_entered"]
    gaps = [g for e in entries if (g := _extract_gap(e.get("reason", ""))) is not None]

    total_closes = len(records)
    stop_loss_closes = [r for r in records if r["close_category"] == "stop_loss"]
    tp_closes = [r for r in records if r["close_category"] == "take_profit"]
    swap_closes = [r for r in records if r["close_category"] == "swapped_confirmed_reversal"]

    swap_pending = [e for e in decisions if e.get("action") == "swap_pending"]
    swap_blocked = [e for e in decisions if e.get("action") == "swap_blocked_low_adx"]
    swap_cancelled = [
        e for e in decisions if e.get("action") == "pending_cancelled" and e.get("reason", "").startswith("Pending ")
    ]
    gap_setup_cancelled = [
        e for e in decisions if e.get("action") == "pending_cancelled" and not e.get("reason", "").startswith("Pending ")
    ]

    session_blocked = [
        e for e in decisions if e.get("action") in ("cross_ignored_outside_session", "pending_touch_outside_session")
    ]
    session_gap_counts = Counter(_session_gap_label(e["_ts"], sessions) for e in session_blocked)

    adx_entry_blocked = [e for e in decisions if e.get("action") == "entry_blocked_adx_falling"]
    validation_failed = [e for e in decisions if e.get("action") == "validation_failed"]

    return {
        "variant": "dual_cross_confirmed_swap_adx",
        "account": account,
        "entries_fired": len(entries),
        "avg_gap": mean(gaps) if gaps else None,
        "gap_histogram": _histogram(gaps),
        "gap_sample_count": len(gaps),
        "entries_no_gap_logged": len(entries) - len(gaps),  # swap re-entries, by design
        "adx_entry_blocked_count": len(adx_entry_blocked),
        "validation_failed_count": len(validation_failed),
        "stop_loss_hits": len(stop_loss_closes),
        "stop_loss_pct": 100 * len(stop_loss_closes) / total_closes if total_closes else 0,
        "take_profit_hits": len(tp_closes),
        "take_profit_pct": 100 * len(tp_closes) / total_closes if total_closes else 0,
        "swap_pending_count": len(swap_pending),
        "swap_confirmed_count": len(swap_closes),
        "swap_blocked_count": len(swap_blocked),
        "swap_cancelled_count": len(swap_cancelled),
        "swap_confirmed_avg_pl": mean([r["profit"] for r in swap_closes]) if swap_closes else None,
        "gap_setup_pending_count": len([e for e in decisions if e.get("action") == "setup_pending"]),
        "gap_setup_cancelled_count": len(gap_setup_cancelled),
        "session_blocked_count": len(session_blocked),
        "session_gap_counts": session_gap_counts,
        "total_closes": total_closes,
    }


def build_rule_tracking_simple_swap(account: str, decisions: list[dict], records: list[dict], sessions: list[SessionWindow]) -> dict:
    """For strategy_variant=dual_cross_confirmed_swap (demo2_m1/demo2_m3):
    NO ADX gate anywhere, NO 2-candle debounce, NO stop-loss tightening --
    the first candle that confirms an opposite cross flips the position
    immediately (see bot/strategy/state_machine_dual_cross_confirmed_swap.py's
    module docstring). There is no "episode" state machine to report on for
    the swap, just a plain count of swapped_reversal closes."""
    entries = [e for e in decisions if e.get("action") == "trade_entered"]
    gaps = [g for e in entries if (g := _extract_gap(e.get("reason", ""))) is not None]

    total_closes = len(records)
    stop_loss_closes = [r for r in records if r["close_category"] == "stop_loss"]
    tp_closes = [r for r in records if r["close_category"] == "take_profit"]
    swap_closes = [r for r in records if r["close_category"] == "swapped_reversal"]

    gap_setup_cancelled = [e for e in decisions if e.get("action") == "pending_cancelled"]

    session_blocked = [
        e for e in decisions if e.get("action") in ("cross_ignored_outside_session", "pending_touch_outside_session")
    ]
    session_gap_counts = Counter(_session_gap_label(e["_ts"], sessions) for e in session_blocked)

    return {
        "variant": "dual_cross_confirmed_swap",
        "account": account,
        "entries_fired": len(entries),
        "avg_gap": mean(gaps) if gaps else None,
        "gap_histogram": _histogram(gaps),
        "gap_sample_count": len(gaps),
        "entries_no_gap_logged": len(entries) - len(gaps),  # swap re-entries, by design
        "stop_loss_hits": len(stop_loss_closes),
        "stop_loss_pct": 100 * len(stop_loss_closes) / total_closes if total_closes else 0,
        "take_profit_hits": len(tp_closes),
        "take_profit_pct": 100 * len(tp_closes) / total_closes if total_closes else 0,
        "swap_reversal_count": len(swap_closes),
        "swap_reversal_pct": 100 * len(swap_closes) / total_closes if total_closes else 0,
        "swap_reversal_avg_pl": mean([r["profit"] for r in swap_closes]) if swap_closes else None,
        "gap_setup_pending_count": len([e for e in decisions if e.get("action") == "setup_pending"]),
        "gap_setup_cancelled_count": len(gap_setup_cancelled),
        "session_blocked_count": len(session_blocked),
        "session_gap_counts": session_gap_counts,
        "total_closes": total_closes,
    }


# -------------------------------------------------------------------------
# Daily aggregation (Tab 7)
# -------------------------------------------------------------------------


def _day_stats(records: list[dict]) -> dict:
    total = len(records)
    wins = sum(1 for r in records if r["outcome"] == "WIN")
    pl = sum(r["profit"] for r in records)
    return {"trades": total, "win_rate": (100 * wins / total) if total else 0.0, "pl": pl}


def build_daily_log(m1_records: list[dict], m3_records: list[dict]) -> list[dict]:
    all_records = m1_records + m3_records
    if not all_records:
        return []
    dates = sorted({trading_day(r["exit_time"]) for r in all_records})
    start, end = dates[0], trading_day(datetime.now(timezone.utc))

    def by_date(records: list[dict], d: date_cls) -> list[dict]:
        return [r for r in records if trading_day(r["exit_time"]) == d]

    rows = []
    d = start
    while d <= end:
        m1_day, m3_day = by_date(m1_records, d), by_date(m3_records, d)
        combined_day = m1_day + m3_day
        m1s, m3s, cs = _day_stats(m1_day), _day_stats(m3_day), _day_stats(combined_day)
        rows.append({
            "date": d,
            "m1_trades": m1s["trades"], "m1_win_rate": m1s["win_rate"], "m1_pl": m1s["pl"],
            "m3_trades": m3s["trades"], "m3_win_rate": m3s["win_rate"], "m3_pl": m3s["pl"],
            "combined_trades": cs["trades"], "combined_win_rate": cs["win_rate"], "combined_pl": cs["pl"],
        })
        d += timedelta(days=1)
    return rows


# -------------------------------------------------------------------------
# Excel styling helpers
# -------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=13, color="1F3864")
KPI_LABEL_FONT = Font(size=10, color="595959")
KPI_VALUE_FONT = Font(bold=True, size=20, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="808080")


def write_header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def write_table(ws, top_row: int, headers: list[str], rows: list[list], start_col: int = 1) -> int:
    write_header_row(ws, top_row, headers, start_col)
    for r_i, row in enumerate(rows):
        for c_i, val in enumerate(row):
            ws.cell(row=top_row + 1 + r_i, column=start_col + c_i, value=val)
    return top_row + 1 + len(rows)


def autofit(ws, max_col: int, max_width: int = 32) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for cell in ws[letter]:
            if cell.value is not None:
                best = max(best, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(best + 2, max_width)


def kpi_cell(ws, row: int, col: int, label: str, value: str) -> None:
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = KPI_LABEL_FONT
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = KPI_VALUE_FONT


# -------------------------------------------------------------------------
# Main
# -------------------------------------------------------------------------


def gather_account_data(account: str, timeframe: str) -> tuple[list[dict], list[dict], dict]:
    decisions = [e for e in read_decisions(account) if e["_ts"] >= TESTING_START_UTC]
    config = load_config(account)

    if decisions:
        earliest = min(e["_ts"] for e in decisions)
    else:
        earliest = TESTING_START_UTC

    connector = MT5Connector(config.mt5)
    connector.connect()
    try:
        # MT5's own deal .time fields (and history_deals_get's query args)
        # use the broker's own time convention, NOT true UTC -- confirmed
        # 2026-08-27 this was missing here, producing entry/exit times off
        # by the broker offset (+3h) throughout M1/M3 Detail. Measure it
        # fresh via the connector, same pattern as
        # scripts/tight_exit_real_trades_report.py.
        offset = mt5_utc_offset(connector, config.symbol)
        mt5_trades = get_closed_trades_range(
            config.symbol, config.execution.magic_number,
            TESTING_START_UTC, datetime.now(timezone.utc), offset,
        )
    finally:
        connector.disconnect()
    # Extra safety: a position that opened before TESTING_START_UTC but
    # closed after it would otherwise sneak in via the range query above
    # (get_closed_trades_range pairs by position, keyed off the exit deal
    # falling inside the range) -- exclude anything whose ENTRY predates
    # the testing-start cutoff too.
    mt5_trades = [t for t in mt5_trades if t["entry_time"].astimezone(timezone.utc) >= TESTING_START_UTC]

    records = build_trade_records(account, timeframe, decisions, mt5_trades)
    sessions = config.sessions.get(config.strategy_variant, [])
    if config.strategy_variant == "dual_cross_confirmed_swap":
        rules = build_rule_tracking_simple_swap(account, decisions, records, sessions)
    else:
        # Default/fallback: dual_cross_confirmed_swap_adx (demo1's lineage)
        # -- also what every account ran before demo2 existed, so this
        # stays the safe default for any variant not explicitly handled
        # above rather than silently mislabeling an unknown one.
        rules = build_rule_tracking_swap_adx(account, decisions, records, sessions)
    return decisions, records, rules


def build_workbook(m1: tuple, m3: tuple) -> Workbook:
    m1_decisions, m1_records, m1_rules = m1
    m3_decisions, m3_records, m3_rules = m3
    all_records = sorted(m1_records + m3_records, key=lambda r: r["exit_time"])

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_summary = wb.create_sheet("Summary")
    ws_m1 = wb.create_sheet("M1 Detail")
    ws_m3 = wb.create_sheet("M3 Detail")
    ws_rules = wb.create_sheet("Strategy Rule Tracking")
    ws_attr = wb.create_sheet("Win-Loss Attribution")
    ws_daily = wb.create_sheet("Daily Log")

    # ---- Daily Log (populated first so Dashboard charts can reference it) ----
    daily_rows = build_daily_log(m1_records, m3_records)
    ws_daily.cell(row=1, column=1, value="Daily Log — one row per calendar day since testing started").font = SECTION_FONT
    headers = [
        "Date", "M1 Trades", "M1 Win Rate %", "M1 P/L", "M3 Trades", "M3 Win Rate %", "M3 P/L",
        "Combined Trades", "Combined Win Rate %", "Combined P/L",
    ]
    row = write_table(ws_daily, 3, headers, [
        [
            r["date"].isoformat(), r["m1_trades"], round(r["m1_win_rate"], 1), round(r["m1_pl"], 2),
            r["m3_trades"], round(r["m3_win_rate"], 1), round(r["m3_pl"], 2),
            r["combined_trades"], round(r["combined_win_rate"], 1), round(r["combined_pl"], 2),
        ]
        for r in daily_rows
    ])
    autofit(ws_daily, len(headers))
    daily_first_data_row, daily_last_data_row = 4, 3 + len(daily_rows)

    # ---- Dashboard ----
    ws_dash.sheet_view.showGridLines = False
    ws_dash.cell(row=1, column=1, value="Live/Shadow Test Report — Dashboard").font = Font(bold=True, size=16, color="1F3864")
    ws_dash.cell(row=2, column=1, value=(
        f"Last updated: {datetime.now(COLOMBO).strftime('%Y-%m-%d %H:%M:%S')} (Asia/Colombo)  |  "
        f"Report date: {date_cls.today().isoformat()}"
    )).font = NOTE_FONT

    total = len(all_records)
    wins = sum(1 for r in all_records if r["outcome"] == "WIN")
    win_rate = (100 * wins / total) if total else 0.0
    total_pl = sum(r["profit"] for r in all_records)

    def streak(records: list[dict]) -> str:
        if not records:
            return "N/A"
        last_outcome = records[-1]["outcome"]
        if last_outcome == "BREAKEVEN":
            return "Breakeven"
        n = 0
        for r in reversed(records):
            if r["outcome"] != last_outcome:
                break
            n += 1
        return f"{n} {'win' if last_outcome == 'WIN' else 'loss'}{'es' if n != 1 and last_outcome == 'LOSS' else 's' if n != 1 else ''}"

    kpi_cell(ws_dash, 4, 1, "Combined Trades", str(total))
    kpi_cell(ws_dash, 4, 3, "Win Rate", f"{win_rate:.1f}%")
    kpi_cell(ws_dash, 4, 5, "Total P/L", f"{total_pl:+.2f}")
    kpi_cell(ws_dash, 4, 7, "Current Streak", streak(all_records))

    ws_dash.cell(row=7, column=1, value="M1 vs M3").font = SECTION_FONT
    m1_total, m3_total = len(m1_records), len(m3_records)
    m1_wins = sum(1 for r in m1_records if r["outcome"] == "WIN")
    m3_wins = sum(1 for r in m3_records if r["outcome"] == "WIN")
    write_table(ws_dash, 8, ["Timeframe", "Trades", "Win Rate %", "P/L"], [
        ["M1", m1_total, round(100 * m1_wins / m1_total, 1) if m1_total else 0, round(sum(r["profit"] for r in m1_records), 2)],
        ["M3", m3_total, round(100 * m3_wins / m3_total, 1) if m3_total else 0, round(sum(r["profit"] for r in m3_records), 2)],
    ])
    autofit(ws_dash, 4)

    today = trading_day(datetime.now(timezone.utc))
    today_records = [r for r in all_records if trading_day(r["exit_time"]) == today]
    if today_records:
        common = Counter(category_label(r["close_category"]) for r in today_records).most_common(1)[0]
        callout = f"{common[0]} ({common[1]} of {len(today_records)} closes today)"
    else:
        callout = "No trades closed today yet"
    ws_dash.cell(row=12, column=1, value="Most Common Close Reason Today").font = SECTION_FONT
    ws_dash.cell(row=13, column=1, value=callout).font = Font(size=12, bold=True)

    if daily_last_data_row >= daily_first_data_row:
        equity_chart = LineChart()
        equity_chart.title = "Combined Equity Curve (Cumulative P/L, Daily)"
        equity_chart.y_axis.title = "Cumulative P/L (USD)"
        equity_chart.x_axis.title = "Date"
        equity_chart.style = 2
        equity_chart.height, equity_chart.width = 9, 18
        # Build a cumulative-P/L helper column on Daily Log (col K) so the
        # chart doesn't need its own separate data block.
        ws_daily.cell(row=3, column=11, value="Combined Cumulative P/L")
        running = 0.0
        for i, r in enumerate(daily_rows):
            running += r["combined_pl"]
            ws_daily.cell(row=4 + i, column=11, value=round(running, 2))
        data = Reference(ws_daily, min_col=11, min_row=3, max_row=daily_last_data_row)
        cats = Reference(ws_daily, min_col=1, min_row=daily_first_data_row, max_row=daily_last_data_row)
        equity_chart.add_data(data, titles_from_data=True)
        equity_chart.set_categories(cats)
        ws_dash.add_chart(equity_chart, "A16")

        pl_chart = BarChart()
        pl_chart.title = "Combined Daily P/L"
        pl_chart.y_axis.title = "P/L (USD)"
        pl_chart.x_axis.title = "Date"
        pl_chart.style = 10
        pl_chart.height, pl_chart.width = 9, 18
        data = Reference(ws_daily, min_col=10, min_row=3, max_row=daily_last_data_row)
        pl_chart.add_data(data, titles_from_data=True)
        pl_chart.set_categories(cats)
        ws_dash.add_chart(pl_chart, "A34")
    else:
        ws_dash.cell(row=16, column=1, value="No closed trades yet — charts will populate once trades exist.").font = NOTE_FONT

    # ---- Summary ----
    def summary_stats(records: list[dict]) -> dict:
        n = len(records)
        w = [r for r in records if r["outcome"] == "WIN"]
        l = [r for r in records if r["outcome"] == "LOSS"]
        return {
            "Total Trades": n, "Wins": len(w), "Losses": len(l),
            "Win Rate %": round(100 * len(w) / n, 1) if n else 0,
            "Total P/L": round(sum(r["profit"] for r in records), 2),
            "Avg Win": round(mean([r["profit"] for r in w]), 2) if w else 0,
            "Avg Loss": round(mean([r["profit"] for r in l]), 2) if l else 0,
        }

    ws_summary.cell(row=1, column=1, value="Summary").font = SECTION_FONT
    combined_stats, m1_stats, m3_stats = summary_stats(all_records), summary_stats(m1_records), summary_stats(m3_records)
    metrics = list(combined_stats.keys())
    write_table(ws_summary, 3, ["Metric", "Combined (M1+M3)", "M1", "M3"], [
        [m, combined_stats[m], m1_stats[m], m3_stats[m]] for m in metrics
    ])
    manual_rejected = sum(1 for e in m1_decisions + m3_decisions if e.get("action") == "manual_trade_rejected")
    ws_summary.cell(row=13, column=1, value=f"Manual/foreign trades auto-rejected (both accounts): {manual_rejected}").font = NOTE_FONT
    autofit(ws_summary, 4)

    # ---- M1 / M3 Detail ----
    def write_detail(ws, records: list[dict], stats: dict, label: str) -> None:
        ws.cell(row=1, column=1, value=f"{label} Detail").font = SECTION_FONT
        write_table(ws, 3, ["Metric", "Value"], [[k, v] for k, v in stats.items()])
        start = 3 + len(stats) + 2
        ws.cell(row=start - 1, column=1, value="Trade-by-Trade").font = SECTION_FONT
        headers = ["Entry Time", "Exit Time", "Direction", "Entry Price", "Exit Price", "P/L", "Close Reason"]
        rows = [
            [
                r["entry_time"].strftime("%Y-%m-%d %H:%M:%S"), r["exit_time"].strftime("%Y-%m-%d %H:%M:%S"),
                r["direction"], r["entry_price"], r["exit_price"], round(r["profit"], 2), category_label(r["close_category"]),
            ]
            for r in records
        ]
        write_table(ws, start, headers, rows)
        autofit(ws, len(headers))

    write_detail(ws_m1, m1_records, m1_stats, "M1")
    write_detail(ws_m3, m3_records, m3_stats, "M3")

    # ---- Strategy Rule Tracking ----
    variant = m1_rules["variant"]  # both legs of one account pair always run the same variant
    ws_rules.cell(row=1, column=1, value=f"Strategy Rule Tracking — {variant}").font = SECTION_FONT
    ws_rules.cell(row=2, column=1, value=(
        "Each account's ACTIVE strategy_variant / sessions are read fresh from its current config at report-run "
        "time — historical entries logged under a different config are still counted by action type, but session-gap "
        "labels reflect today's config, not necessarily what was active when an old event fired."
    )).font = NOTE_FONT

    r = 4
    ws_rules.cell(row=r, column=1, value="a) Entry (gap-tolerance) rule").font = Font(bold=True, size=12)
    r += 1
    entry_rows = [
        ["Entries fired", m1_rules["entries_fired"], m3_rules["entries_fired"]],
        ["Entries with a logged gap (flat entries)", m1_rules["gap_sample_count"], m3_rules["gap_sample_count"]],
        ["Entries with no gap logged (swap re-entries, by design)", m1_rules["entries_no_gap_logged"], m3_rules["entries_no_gap_logged"]],
        ["Avg |close-EMA13| gap at entry ($)",
         round(m1_rules["avg_gap"], 3) if m1_rules["avg_gap"] is not None else "N/A",
         round(m3_rules["avg_gap"], 3) if m3_rules["avg_gap"] is not None else "N/A"],
    ]
    if variant == "dual_cross_confirmed_swap_adx":
        entry_rows.append(
            ["ADX-momentum filter blocks (bonus filter, not gap-related)", m1_rules["adx_entry_blocked_count"], m3_rules["adx_entry_blocked_count"]]
        )
    r = write_table(ws_rules, r, ["Metric", "M1", "M3"], entry_rows)
    ws_rules.cell(row=r, column=1, value="Gap size histogram (M1)").font = Font(italic=True)
    r += 1
    r = write_table(ws_rules, r, ["Bucket", "Count"], [[b, c] for b, c in m1_rules["gap_histogram"]])
    r += 1
    ws_rules.cell(row=r, column=1, value="Gap size histogram (M3)").font = Font(italic=True)
    r += 1
    r = write_table(ws_rules, r, ["Bucket", "Count"], [[b, c] for b, c in m3_rules["gap_histogram"]])
    r += 2

    ws_rules.cell(row=r, column=1, value="b) Validation-at-close rule").font = Font(bold=True, size=12)
    r += 1
    ws_rules.cell(row=r, column=1, value=(
        f"N/A for {variant} — every position opens already close-confirmed, so there is no own-candle "
        "validation step and no validation_failed category (see the engine's module docstring)."
    )).font = NOTE_FONT
    r += 2

    ws_rules.cell(row=r, column=1, value="c) Stop-loss rule").font = Font(bold=True, size=12)
    r += 1
    if variant != "dual_cross_confirmed_swap_adx":
        ws_rules.cell(row=r, column=1, value=(
            "No pending-reversal stop-loss tightening in this variant — the stop-loss stays at its full "
            "configured distance for the whole life of every position (see the engine's module docstring)."
        )).font = NOTE_FONT
        r += 1
    r = write_table(ws_rules, r, ["Metric", "M1", "M3"], [
        ["Stop-loss hits", m1_rules["stop_loss_hits"], m3_rules["stop_loss_hits"]],
        ["% of all closes", round(m1_rules["stop_loss_pct"], 1), round(m3_rules["stop_loss_pct"], 1)],
    ])
    r += 2

    ws_rules.cell(row=r, column=1, value="d) Take-profit rule").font = Font(bold=True, size=12)
    r += 1
    r = write_table(ws_rules, r, ["Metric", "M1", "M3"], [
        ["Take-profit hits", m1_rules["take_profit_hits"], m3_rules["take_profit_hits"]],
        ["% of all closes", round(m1_rules["take_profit_pct"], 1), round(m3_rules["take_profit_pct"], 1)],
    ])
    r += 2

    ws_rules.cell(row=r, column=1, value="e) Swap / reversal rule").font = Font(bold=True, size=12)
    r += 1
    if variant == "dual_cross_confirmed_swap_adx":
        ws_rules.cell(row=r, column=1, value=(
            "dual_cross_confirmed_swap_adx holds at most ONE position at a time (single-position, reversal-based "
            "engine) — there is no scenario where a 2nd position opens while the 1st is still open, so this is "
            "reframed as: episode = the FIRST opposing candle arming a pending reversal (swap_pending) while a "
            "position is held. Each episode resolves one of three ways: CONFIRMED (2nd candle + ADX both pass -> "
            "old position closes, new one opens immediately), BLOCKED_BY_ADX (2nd candle confirms but ADX too weak "
            "-> old position keeps running, stop-loss stays tightened), or CANCELLED (no 2nd opposing candle -> "
            "old position keeps running, tightened stop stays as-is)."
        )).font = NOTE_FONT
        r += 1
        r = write_table(ws_rules, r, ["Outcome", "M1", "M3"], [
            ["Episodes armed (swap_pending)", m1_rules["swap_pending_count"], m3_rules["swap_pending_count"]],
            ["Confirmed (closed original, opened new)", m1_rules["swap_confirmed_count"], m3_rules["swap_confirmed_count"]],
            ["Blocked by ADX (no close)", m1_rules["swap_blocked_count"], m3_rules["swap_blocked_count"]],
            ["Cancelled (no 2nd candle, no close)", m1_rules["swap_cancelled_count"], m3_rules["swap_cancelled_count"]],
            ["Avg P/L of the closed trade on CONFIRMED outcomes",
             round(m1_rules["swap_confirmed_avg_pl"], 2) if m1_rules["swap_confirmed_avg_pl"] is not None else "N/A",
             round(m3_rules["swap_confirmed_avg_pl"], 2) if m3_rules["swap_confirmed_avg_pl"] is not None else "N/A"],
        ])
    else:
        ws_rules.cell(row=r, column=1, value=(
            "dual_cross_confirmed_swap has NO debounce and NO ADX gate on the swap — the very first candle "
            "whose close confirms an opposite EMA13/21 cross closes the held position and opens the new one "
            "the same instant, regardless of P/L (category swapped_reversal — deliberately a different name "
            "from demo1's swapped_confirmed_reversal, since the semantics genuinely differ)."
        )).font = NOTE_FONT
        r += 1
        r = write_table(ws_rules, r, ["Metric", "M1", "M3"], [
            ["Swap reversal closes", m1_rules["swap_reversal_count"], m3_rules["swap_reversal_count"]],
            ["% of all closes", round(m1_rules["swap_reversal_pct"], 1), round(m3_rules["swap_reversal_pct"], 1)],
            ["Avg P/L of swap-reversal closes",
             round(m1_rules["swap_reversal_avg_pl"], 2) if m1_rules["swap_reversal_avg_pl"] is not None else "N/A",
             round(m3_rules["swap_reversal_avg_pl"], 2) if m3_rules["swap_reversal_avg_pl"] is not None else "N/A"],
        ])
    r += 2

    ws_rules.cell(row=r, column=1, value="f) 2-position cap rule").font = Font(bold=True, size=12)
    r += 1
    ws_rules.cell(row=r, column=1, value=(
        f"N/A for {variant} — this engine has no concurrent-position code path at all (self.position is a "
        "single optional slot, never a list); a 2-position cap only exists in the older dual_cross variant "
        "(state_machine_dual_cross.py's max_concurrent_positions), which is not what's live on this account."
    )).font = NOTE_FONT
    r += 2

    ws_rules.cell(row=r, column=1, value="g) Session window rule").font = Font(bold=True, size=12)
    r += 1
    r = write_table(ws_rules, r, ["Metric", "M1", "M3"], [
        ["Blocked-by-session events", m1_rules["session_blocked_count"], m3_rules["session_blocked_count"]],
    ])
    r += 1
    ws_rules.cell(row=r, column=1, value="Blocked events by inter-window gap (M1)").font = Font(italic=True)
    r += 1
    r = write_table(ws_rules, r, ["Gap", "Count"], [[k, v] for k, v in m1_rules["session_gap_counts"].most_common()])
    r += 1
    ws_rules.cell(row=r, column=1, value="Blocked events by inter-window gap (M3)").font = Font(italic=True)
    r += 1
    r = write_table(ws_rules, r, ["Gap", "Count"], [[k, v] for k, v in m3_rules["session_gap_counts"].most_common()])
    autofit(ws_rules, 3)

    # ---- Win-Loss Attribution ----
    ws_attr.cell(row=1, column=1, value="Win/Loss Attribution — every closed trade, M1+M3 combined").font = SECTION_FONT
    headers = ["Date", "Timeframe", "Direction", "Entry Price", "Exit Price", "P/L", "Outcome", "Close Reason", "Entry Gap ($)"]
    rows = [
        [
            r["exit_time"].strftime("%Y-%m-%d"), r["timeframe"], r["direction"], r["entry_price"], r["exit_price"],
            round(r["profit"], 2), r["outcome"], category_label(r["close_category"]),
            round(r["entry_gap"], 2) if r["entry_gap"] is not None else "N/A",
        ]
        for r in all_records
    ]
    write_table(ws_attr, 3, headers, rows)
    autofit(ws_attr, len(headers))

    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--m1-account", default="demo1_m1")
    parser.add_argument("--m3-account", default="demo1_m3")
    parser.add_argument(
        "--output-name", default="live_test_report.xlsx",
        help="Filename under reports/live_test/ (default matches the original demo1 report -- "
             "pass a distinct name, e.g. demo2_report.xlsx, for a different account pair so it "
             "doesn't overwrite demo1's file).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    print(f"Reading decisions + MT5 history for {args.m1_account} (M1)...")
    m1 = gather_account_data(args.m1_account, "M1")
    print(f"Reading decisions + MT5 history for {args.m3_account} (M3)...")
    m3 = gather_account_data(args.m3_account, "M3")

    print("Building workbook...")
    wb = build_workbook(m1, m3)

    out_path = PROJECT_ROOT / "reports" / "live_test" / args.output_name
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    m1_records, m3_records = m1[1], m3[1]
    print(f"Written: {out_path}")
    print(f"M1: {len(m1_records)} trades | M3: {len(m3_records)} trades | Combined: {len(m1_records) + len(m3_records)} trades")


if __name__ == "__main__":
    main()
